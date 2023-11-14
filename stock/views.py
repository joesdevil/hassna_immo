import csv
import os
import openpyxl
from django.conf import settings
from django.contrib import messages
from django.db.models import Q
from django.db.models import Count
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import *
from .form import *
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.http import JsonResponse
from django.db.models.functions import TruncDate

# Create your views here.


def new_register(request):
    form = UserCreationForm
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/account/login/')
    context = {'form': form}
    return render(request, 'stock/register.html', context)



def get_client_ip(request):
    
    if request.user.is_authenticated:

        labels = []
        label_item = []
        label_item1 = []
        label_item11 = []
        label_item_b = []
        label_item_p = []
        reservation=AddTask.objects.all()
        data_b=[]
        data_price=[]
        label_price=[]
        data = []
        data1 = []
        data11 = []
        data_p = []
        issue_data = []
        receive_data = []
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        u = User(user=ip)
        result = User.objects.filter(Q(user__icontains=ip))
        if len(result) == 1:
            pass
        else:
            u.save()
            return ip
        queryset = Stock.objects.all()
        querysetProj=Project.objects.all()
        querys = Category.objects.all()
        querysetpart=AddTask.objects.all()
        
        queryset_task=AddTask.objects.filter(confirmed=True)
        for chart in queryset_task:
            label_price.append(str(chart.dateReservation.date()) )
            data_price.append(chart.deposit_amount)
        print("date price: ",data_price)


        for chart in queryset:
            label_item.append(chart.nomProject)
            data.append(chart.superficieHabitable)
            issue_data.append(chart.superficieHabitable)
            receive_data.append(chart.superficieHabitable)

        for chart in querysetpart:
            label_item11.append(chart.customer.nom)
            data11.append(chart.remaining_parts)

        
        label_item_p.append(len(queryset.filter(etat="Libre")))
        data_p.append(len(queryset.filter(etat="Libre")))

        label_item_p.append(len(queryset.filter(etat="Réservé")))
        data_p.append(len(queryset.filter(etat="Réservé")))

        label_item_p.append(len(queryset.filter(etat="Vendu")))
        data_p.append(len(queryset.filter(etat="Vendu")))
        
        label_item_p=["Libre", "Réservé", "Vendu"]
            
        


        for chart in querysetProj:
            label_item1.append(chart.nom)
            data1.append(chart.nbrLotsTotal)
            

        for chart in querys:
            labels.append(str(chart.group))

        today = datetime.datetime.now().date()
        task_count_by_day = AddTask.objects.annotate(day=TruncDate('dateReservation')).values('day').annotate(count=Count('id')).order_by('day')
        # print("task: ",task_count_by_day )
        for task_count in task_count_by_day :
            data_b.append(task_count["count"])
            year = task_count["day"].year
            month = task_count["day"].month
            day_of_month = task_count["day"].day
            label_item_b.append( f"{year}-{month:02d}-{day_of_month:02d}")
    
        
            
        count = Stock.objects.all().count()
        body = Project.objects.all().count()
        mind = AddTask.objects.all().count()
        soul = Person.objects.all().count()
        reservation=AddTask.objects.all()
        context = {
            'count': count,
            'body': body,
            'mind': mind,
            'soul': soul,
            'labels': labels,
            'data': data,
            'data1': data1,
            'data11': data11,
            'data_b': data_b,
            'issue_data': issue_data,
            'receive_data': receive_data,
            'label_item': label_item,
            'label_item1': label_item1,
            'label_item11': label_item11,
            'label_item_b': label_item_b,
            'data_price':data_price,
            'data_p':data_p,
            'label_item_p': label_item_p,
            'reservation':reservation,
            'label_price':label_price,
            'present_date': datetime.date.today()
        }
        return render(request, 'stock/home.html', context)
    else:
        return redirect("/account/login/")

import datetime
@login_required
def view_stock(request):
    reservation=AddTask.objects.all()
    title = "VIEW STOCKS"
    everything = Stock.objects.all()
    try:
        type=request.GET["typeBien"]
        etat=request.GET["etat"]
        if type and etat:
            everything = everything.filter(typeBien=type,etat=etat)
        elif type:
            everything = everything.filter(typeBien=type)
        else :
            everything = everything.filter(etat=False)
                 
    except:
        pass
    form = StockSearchForm(request.POST or None)
    present_date= datetime.date.today()
    five_days_ahead= present_date + datetime.timedelta(days=5)
    context = {'everything': everything,'reservation':reservation, 'form': form,'present_date':present_date,'five_days_ahead':five_days_ahead}
    if request.method == 'POST':
        print("posted post")
        # type = form['type'].value()
        # everything = Stock.objects.filter(typeBien = type)
        # if type != '':
        #     everything = everything.filter(category_id=type)

        if form['export_to_CSV'].value() == True:
            instance = everything
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["ID",'Nom', 'BLOCK','ETAGE','COTE','Type Bien',"superficie Habitable", 'prixVenteM2','montant Vente Total','Etat'])

            for stock in instance:
                ws.append([stock.id, stock.nomProject,stock.bloc,stock.etage,stock.cote,stock.typeBien,stock.typeBien, stock.prixVenteM2,stock.montantVenteTotal,stock.etat])

            # Prepare the response for Excel download
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Invoice.xlsx'
            wb.save(response)
            return response
    
        context = {'title': title,'present_date': datetime.date.today(), 'everything': everything,'reservation':reservation, 'form': form,'present_date':present_date,'five_days_ahead':five_days_ahead}
    return render(request, 'stock/view_stock.html', context)


@login_required
def scrum_list(request):
    title = 'Add List'
    add = ScrumTitles.objects.all()
    sub = Scrums.objects.all()
    if request.method == 'POST':
        form = AddScrumListForm(request.POST, prefix='banned')
        if form.is_valid():
            form.save()
    else:
        form = AddScrumListForm(prefix='banned')
    if request.method == 'POST' and not form:
        task_form = AddScrumTaskForm(request.POST, prefix='expected')
        form = AddScrumListForm(prefix='banned')
        if task_form.is_valid():
            task_form.save()
    else:
        task_form = AddScrumTaskForm(prefix='expected')
    context = {'add': add, 'form': form, 'task_form': task_form, 'sub': sub, 'title': title}
    return render(request, 'stock/scrumboard.html', context)





@login_required
def scrum_view(request):
    title = 'View'
    viewers = ScrumTitles.objects.all()
    context = {'title': title, 'view': viewers}
    return render(request, 'stock/scrumboard.html', context)


@login_required
def add_stock(request):
    reservation=AddTask.objects.all()
    print("hmmmm")
    title = 'Add Reservation'
    add = Stock.objects.all()
    
    if request.method == 'POST':
        form = StockCreateForm(request.POST, request.FILES)
        if form.is_valid():
            print("saved post")
            form.save()
            messages.success(request, 'Successful')
            return redirect('/view_stock')
    else:
        form = StockCreateForm()
        print("get res")
    context = {'add': add, 'form': form,'present_date': datetime.date.today(),'reservation':reservation, 'title': title}
     
    return render(request, 'stock/add_stock.html', context)

@login_required
def add_cat(request):
     
    title = 'Add Category'
    add = Category.objects.filter(user=request.user)
    
    if request.method == 'POST':
        form = CategoryCreateForm(request.POST)
        if form.is_valid():
            Category.objects.get_or_create(
                group=request.POST["group"],
                user=request.user
            )
            
            messages.success(request, 'Successful')
            return redirect('/add_stock')
    else:
        form = CategoryCreateForm()
        print("get res")
    context = {'add': add, 'form': form, 'title': title}
     
    return render(request, 'stock/add-cat.html', context)

@login_required
def add_loc(request):
     
    title = 'Add Country'
    add = Category.objects.filter(user=request.user)
    
    if request.method == 'POST':
        form = AddCountry(request.POST)
        
        if form.is_valid()  :
            Country.objects.get_or_create(
                name=form.cleaned_data["name"],
                 
            )
            return redirect(f'/add_loc1?q={form.cleaned_data["name"]}')
          
            
        messages.success(request, 'Successful')
        return redirect('/add_loc1')
    else:
        form = AddCountry(request.POST)
         
        print("get res")
    context = {'add': add, 'form': form,   'title': title}
     
    return render(request, 'stock/add_loc.html', context)

@login_required
def add_loc1(request):
     
    title = 'Add State'
    add = Category.objects.filter(user=request.user)
    
    if request.method == 'POST':
        country= request.GET["q"]
        country_fil=Country.objects.filter(name=country).first()
      
        form = AddState(request.POST)
       
   
           
        if form.is_valid()  :
            State.objects.get_or_create(
                country_id=country_fil.id,
                name=form.cleaned_data["name"]
                
            )
            return redirect(f'/add_loc2?q={form.cleaned_data["name"]}')
        
       
        
        messages.success(request, 'Successful')
        return redirect('/add_loc2')
    else:
         
        form = AddState(request.POST)
         
      
    context = {'add': add,   'form': form, 'title': title}
     
    return render(request, 'stock/add_loc1.html', context)


@login_required
def add_loc2(request):
    reservation=AddTask.objects.all()
     
    title = 'Add City'
    add = Category.objects.filter(user=request.user)
    state= request.GET["q"]
    state_fil=State.objects.filter(name=state).first()
    if request.method == 'POST':
      
        form = AddCity(request.POST)
       
   
           
        if form.is_valid()  :
            City.objects.get_or_create(
                state_id=state_fil.id,
                name=form.cleaned_data["name"]
                
            )
            return redirect('/dependent_forms')
        
       
        
        messages.success(request, 'Successful')
        return redirect('/dependent_forms')
    else:
         
        form = AddCity(request.POST)
         
      
    context = {'add': add,   'form': form, 'title': title}
     
    return render(request, 'stock/add_loc.html', context)

@login_required
def update_stock(request, pk):
    reservation=AddTask.objects.all()
    title = 'Update Stock'
    update = Stock.objects.get(id=pk)
    form = StockUpdateForm(instance=update)
    if request.method == 'POST':
        form = StockUpdateForm(request.POST, request.FILES, instance=update)
        if form.is_valid():
            # image_path = update.image.path
            # if os.path.exists(image_path):
                # os.remove(image_path)
            form.save()
            messages.success(request, 'Successfully Updated!')
            return redirect('/view_stock')
    context = {'form': form,'present_date': datetime.date.today(), 'update': update,'reservation':reservation, 'title': title}
    return render(request, 'stock/add_stock.html', context)


@login_required
def update_project(request, pk):
    reservation=AddTask.objects.all()
    title = 'Update Project'
    update = Project.objects.get(id=pk)
    form = ProjectUpdateForm(instance=update)
    if request.method == 'POST':
        form = ProjectUpdateForm(request.POST , instance=update)
        if form.is_valid():
             
            form.save()
            messages.success(request, 'Successfully Updated!')
            return redirect('/view_project')
    context = {'form': form, 'update': update,'present_date': datetime.date.today(), 'reservation':reservation,'title': title}
    return render(request, 'stock/add_project.html', context)



@login_required
def update_task(request, pk):
    reservation=AddTask.objects.all()
    title = 'Update Task'
    update = AddTask.objects.get(id=pk)
    form = TaskUpdateForm(instance=update)
    if request.method == 'POST':
        form = TaskUpdateForm(request.POST,  instance=update)
        if form.is_valid():
            
            # update.user.id=form['user'].value()
            form.save()

            
            messages.success(request, 'Successfully Updated!')
            return redirect('/task_view')
    context = {'form': form,'present_date': datetime.date.today(), 'update': update, 'reservation':reservation,'title': title}
    return render(request, 'stock/add_task.html', context)


@login_required
def delete_stock(request, pk):
    Stock.objects.get(id=pk).delete()
    messages.success(request, 'Your file has been deleted.')
    return redirect('/view_stock')


@login_required
def delete_task(request, pk):
    AddTask.objects.get(id=pk).delete()
    messages.success(request, 'Your task has been deleted.')
    return redirect('/task_view')


@login_required
def delete_project(request, pk):
    Project.objects.get(id=pk).delete()
    messages.success(request, 'Your Project has been deleted.')
    return redirect('/view_project')

@login_required
def stock_detail(request, pk):
    reservation=AddTask.objects.all()
    detail = Stock.objects.get(id=pk)
    context = {
        'detail': detail,'reservation':reservation,
    }
    return render(request, 'stock/stock_detail.html', context)


@login_required
def issue_item(request, pk):
    reservation=AddTask.objects.all()
    issue = Stock.objects.get(id=pk)
    form = IssueForm(request.POST or None, instance=issue)
    if form.is_valid():
        value = form.save(commit=False)
        value.receive_quantity = 0
        value.quantity = value.quantity - value.issue_quantity
        value.issued_by = str(request.user)
        if value.quantity >= 0:
            messages.success(request, "Issued Successfully, " + str(value.quantity) + " " + str(
                value.item_name) + "s now left in Store")
            value.save()
        else:
            messages.error(request, "Insufficient Stock")

        return redirect('/stock_detail/' + str(value.id))
     
        
    context = {
        "title": 'Issue ' + str(issue.item_name),
        "issue": issue,
        "form": form,
        "username": 'Issued by: ' + str(request.user),
    }
    return render(request, "stock/add_stock.html", context)


@login_required
def receive_item(request, pk):
    reservation=AddTask.objects.all()
    receive = Stock.objects.get(id=pk)
    form = ReceiveForm(request.POST or None, instance=receive)
    if form.is_valid():
        value = form.save(commit=False)
        value.issue_quantity = 0
        value.quantity = value.quantity + value.receive_quantity
        value.received_by = str(request.user)
        value.save()
        messages.success(request, "Received Successfully, " + str(value.quantity) + " " + str(
            value.item_name) + "s now in Store")

        return redirect('/stock_detail/' + str(value.id))
    context = {
        "title": 'Receive ' + str(receive.item_name),
        "receive": receive,
        "form": form,
        "username": 'Received by: ' + str(request.user),
    }
    return render(request, "stock/add_stock.html", context)

@login_required
def co_order(request, pk):
    reservation=AddTask.objects.all()
    order = AddTask.objects.get(id=pk)
    order.confirmed = True
    order.save()
    messages.success(request, 'Confirmed order')
    return redirect('/task_view')
    

@login_required
def re_order(request, pk):
    reservation=AddTask.objects.all()
    order = Stock.objects.get(id=pk)
    form = ReorderLevelForm(request.POST or None, instance=order)
    if form.is_valid():
        value = form.save(commit=False)
        value.save()
        messages.success(request, 'Reorder level for ' + str(value.item_name) + ' is updated to ' + str(value.re_order))
        return redirect('/view_stock')
    context = {
        'value': order,
        'form': form
    }
    return render(request, 'stock/add_stock.html', context)


@login_required()
def view_history(request):
    reservation=AddTask.objects.all()
    title = "STOCK HISTORY"
    history = StockHistory.objects.all()
    form = StockHistorySearchForm(request.POST or None)
    context = {
        'title': title,
        'history': history,
        'form': form
    }
    if request.method == 'POST':
        category = form['category'].value()
       
        history = StockHistory.objects.filter(item_name__icontains=form['item_name'].value())
        if category != '':
            history = history.filter(category_id=category)

        if form['export_to_CSV'].value() == True:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="Stock History.csv"'
            writer = csv.writer(response)
            writer.writerow(
                ['CATEGORY',
                 'ITEM NAME',
                 'QUANTITY',
                 'ISSUE QUANTITY',
                 'RECEIVE QUANTITY',
                 'RECEIVE BY',
                 'ISSUE BY',
                 'LAST UPDATED'])
            instance = history
            for stock in instance:
                writer.writerow(
                    [stock.category,
                     stock.item_name,
                     stock.quantity,
                     stock.issue_quantity,
                     stock.receive_quantity,
                     stock.received_by,
                     stock.issued_by,
                     stock.last_updated])
            return response
        context = {
            'form': form,
            'title': title,
            'history': history,
        }
    return render(request, 'stock/view_history.html', context)


@login_required
def dependent_forms(request):
    reservation=AddTask.objects.all()
    title = 'Dependent Forms'
    form = DependentDropdownForm()
    if request.method == 'POST':
        form = DependentDropdownForm(request.POST )
        if form.is_valid():
            form.save()
             
             
            messages.success(request,   ' Successfully Added!')
            return redirect('/depend_form_view')
    context = {'form': form,'present_date': datetime.date.today(), 'reservation':reservation,'title': title}
    return render(request, 'stock/add_dep.html', context)


@login_required
def add_task(request):
    reservation=AddTask.objects.all()
    title = 'RESERVATION'
    form = AddTaskForm()
    if request.method == 'POST':
        form = AddTaskForm(request.POST )
        if form.is_valid():
            print("form s valid"   )
            updatestock=Stock.objects.filter(pk=request.POST["product"]).first()
            print("query is ", updatestock.etat)
            if updatestock.etat == "Libre":
                print("updatestock.etat ==  Libre ")
                confirmed_ord=False
                if form["deposit_amount"] == updatestock.montantVenteTotal:
                    print("updatestock.etat ==  Vendu ")
               
                    updatestock.product.etat  = "Vendu" 
                    confirmed_ord=True
                # elif form["deposit_amount"] > form["total_amount"]:
                
                #     updatestock.etat  = "Libre"
                else:
                    updatestock.etat  = "Réservé"
                    confirmed_ord=True
                    print("updatestock.etat == Réservé")

                # updatestock.product.save()
                updatestock.save()
                print("saving stock")
               
                
                
                AddTask.objects.get_or_create(
                    customer_id=request.POST["customer"], 
                    # user = request.user,
                    product_id=updatestock.id,
                    deposit_amount=request.POST["deposit_amount"],
                    total_amount=updatestock.montantVenteTotal,
                    parts=request.POST["parts"],
                    dateReservation=request.POST["dateReservation"],
                    confirmed=confirmed_ord
                )
                
                print("saving AddTask")
                messages.success(request, str(form['customer'].value()) + f' à été {updatestock.etat} avec success!')
                return redirect('/task_view')
            
            else:
                 
                messages.error(request, str(form['customer'].value()) +  ' est déja ' +updatestock.etat)
                
                return redirect('/add_task')
    context = {'form': form, 'reservation':reservation,'present_date': datetime.date.today(),'title': title}
    return render(request, 'stock/add_task.html', context)

@login_required
def task_view(request):
    reservation=AddTask.objects.all()
    title = "VIEW STOCKS"
    everything = AddTask.objects.all()
    form = StockSearchForm(request.POST or None)
    
    try:
        type=request.GET["typeBien"]
        etat=request.GET["etat"]
        if type and etat:
            everything = everything.filter(typeBien=type,etat=etat)
        elif type:
            everything = everything.filter(typeBien=type)
        else :
            everything = everything.filter(etat=False)
                 
    except:
        pass
        
        
    present_date= datetime.date.today()
    context = {'everything': everything,'reservation':reservation, 'form': form,"present_date":present_date}
    if request.method == 'POST':
        print("posted post")
        # category = form['category'].value()
        # everything = AddTask.objects.all()
        # if category != '':
        #     everything = everything.filter(category_id=category)

        if form['export_to_CSV'].value() == True:
            instance = everything
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['CUSTOMER', 'PROJECT','BIEN','BLOCK','ETAGE','COTE',"superficie Habitable","prix Vente M2","montant Vente Total","etat",    "parts","parts remains" ,  'TOTAL AMOUNT','DEPOSIT AMOUNT',"PAYEMENT TYPE", 'PHONE NUMBER', "DATE","confirmed"])

            for instance1 in instance:
                print("length is", len(instance))
                ws.append([str(instance1.customer), str(instance1.product),str(instance1.product),str(instance1.product.bloc),str(instance1.product.etage),
                           str(instance1.product.cote),instance1.product.superficieHabitable,instance1.product.prixVenteM2,instance1.product.montantVenteTotal,
                           instance1.product.etat,instance1.parts,instance1.parts,  instance1.total_amount,instance1.deposit_amount,instance1.payement_type,
                           instance1.customer.phone,str(instance1.dateReservation.date()),instance1.confirmed ])

            # Prepare the response for Excel download
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Invoice.xlsx'
            wb.save(response)
            return response
        
        context = {'title': title,'present_date': datetime.date.today(), 'reservation':reservation,'everything': everything, 'form': form,"present_date":present_date}
    return render(request, 'stock/task_view.html', context)


@login_required
def dependent_forms_update(request, pk):
    reservation=AddTask.objects.all()
    
    title = 'Update Form'
    dependent_update = Person.objects.get(id=pk)
    form = DependentDropdownForm(instance=dependent_update)
    if request.method == 'POST':
        form = DependentDropdownForm(request.POST, instance=dependent_update)
        if form.is_valid():
            form.save()
            messages.success(request, 'Successfully Updated!')
            return redirect('/depend_form_view')
    context = {
        'title': title,
        'dependent_update': dependent_update,
        'reservation':reservation,
        'present_date': datetime.date.today(),
        'form': form
    }
    return render(request, 'stock/add_dep.html', context)


@login_required
def dependent_forms_view(request):
    reservation=AddTask.objects.all()
    title = 'Dependent Views'
    # viewers = Person.objects.filter(user=request.user)
    viewers=Person.objects.all()
    form = DepSearchForm()
    
    if request.method =='POST':
        form=DepSearchForm(request.POST)
        if form.is_valid():
            # user = form['user'].value()
            everything = Person.objects.all()
            
            if form['export_to_CSV'].value() == True:
                try:
                    current=Person.objects.all()
                    everything = everything.all()
                except:
                    messages.error(request, 'User do not exist')
                    return redirect('/depend_form_view')
                    
            if form['export_to_CSV'].value() == True:
                instance = everything
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(['ID', 'NOM','date Naissance  ','lieu Naissance','phone'
                           ,'email','date Dossier'])

                for task in instance:
                    ws.append([task.id, task.nom,str(task.dateNaissance.date()), 
                               task.lieuNaissance,task.phone,task.email,
                               str(task.dateDossier.date()) ])

                # Prepare the response for Excel download
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=Invoice.xlsx'
                wb.save(response)
                return response
    
            
    context = {'title': title,'present_date': datetime.date.today(), 'reservation':reservation,'view': viewers,'form':form}
    return render(request, 'stock/depend_form_view.html', context)


 
def delete_dependant(request, pk):
    reservation=AddTask.objects.all()
    Person.objects.get(id=pk).delete()
    messages.success(request, 'Your file has been deleted.')
    return redirect('/depend_form_view')


def load_stats(request):
    country_idm = request.GET.get('country_id')
    states = State.objects.filter(country_id=country_idm)
    context = {'states': states}
    return render(request, 'stock/state_dropdown_list_options.html', context)


def load_cities(request):
    reservation=AddTask.objects.all()
    state_main_id = request.GET.get('state_id')
    cities = City.objects.filter(state_id=state_main_id)
    context = {'cities': cities}
    return render(request, 'stock/city_dropdown_list_options.html', context)


@login_required
def contact(request):
    reservation=AddTask.objects.all()
    title = 'Contacts'
    people = Contacts.objects.all()
    form = ContactsForm(request.POST or None)
    if request.method == 'POST':
        form = ContactsForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Successfully Added')
            return redirect('/contacts')
    context = {'people': people,'present_date': datetime.date.today(),'reservation':reservation, 'form': form, 'title': title}
    return render(request, 'stock/contacts.html', context)


@login_required
def addProject(request):
    title = 'Add project'
    reservation=AddTask.objects.all()
    form = AddProject(request.POST or None)
    if request.method == 'POST':
            form = AddProject(request.POST )
            if form.is_valid():
                form.save()
                messages.success(request, 'Successfully Added')
                return redirect('/view_project')
    
    context={"form":form,'present_date': datetime.date.today(),'reservation':reservation,'reservation':reservation}
    return render(request, 'stock/add_project.html', context)

@login_required
def view_project(request):
    title = 'view project'
    everything = Project.objects.all()
    reservation=AddTask.objects.all()
    present_date= datetime.date.today()
   
    context={
        'everything': everything,
        'reservation':reservation,
        'present_date':present_date
    }
    return render(request, 'stock/project_view.html', context)

from django.contrib.auth import logout
def logout_view(request):
    logout(request)
    return redirect('/account/login') 


from django.contrib.auth.views import LoginView
class CustomLoginView(LoginView):
    template_name = 'stock/login.html'